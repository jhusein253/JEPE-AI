import streamlit as st
import openpyxl
import io
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter

# Setup halaman
st.set_page_config(page_title="JEPE AI Pro", layout="wide")
st.title("JEPE AI - Advanced Scanner")

# --- FUNGSI HELPER ---
def clean_int(v):
    try: return int(float(str(v).strip()))
    except (ValueError, TypeError): return None

def generate_excel(original_ws, highlighted_data):
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    for col_num in range(1, original_ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        new_ws.column_dimensions[col_letter].width = 3
    
    colors = {0: "3399FF", 1: "D2B48C", 2: "22C55E", 3: "FFD700"}
    
    for r in range(1, original_ws.max_row + 1):
        for c in range(1, original_ws.max_column + 1):
            cell_val = original_ws.cell(row=r, column=c).value
            new_ws.cell(row=r, column=c).value = cell_val
            
            if (r, c) in highlighted_data:
                pos = highlighted_data[(r, c)]["pos"]
                hex_color = colors.get(pos, "FFFF00")
                new_ws.cell(row=r, column=c).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    
    buf = io.BytesIO()
    new_wb.save(buf)
    buf.seek(0)
    return buf

def render_grid(ws, start_cols, hari_tabel, highlighted, prediction_cells):
    html = ["<div style='overflow-x: auto;'><table style='border-collapse: collapse; width: 100%; text-align: center; font-family: monospace; font-size: 12px;'>"]
    html.append("<tr style='background-color: #0f172a; color: white;'><th>Line</th>")
    for h in hari_tabel:
        html.append(f"<th colspan='4'>{h}</th><th style='width: 15px;'></th>") 
    html.append("</tr>")
    
    for r in range(max(1, ws.max_row - 30), ws.max_row + 1):
        html.append(f"<tr><td style='border: 1px solid #ccc; background-color: #f0f0f0; width: 25px; height: 25px; font-weight: bold;'>{r}</td>")
        for i, start_col in enumerate(start_cols):
            for offset in range(4):
                c_idx = start_col + offset
                val = ws.cell(row=r, column=c_idx).value
                display_val = str(val) if val is not None else "-"
                
                is_pred = (r, c_idx) in prediction_cells
                bg = "#ffffff"
                if (r, c_idx) in highlighted:
                    p = highlighted[(r, c_idx)]["pos"]
                    colors = {0: "#3399FF", 1: "#D2B48C", 2: "#22C55E", 3: "#FFD700"}
                    bg = colors.get(p, "#ffffff")
                elif is_pred: bg = "#fee2e2" 
                
                border_style = "2px solid #dc2626" if is_pred else "1px solid #ccc"
                text_color = "#dc2626" if is_pred else "inherit"
                html.append(f"<td style='border: {border_style}; background-color: {bg}; color: {text_color}; font-weight: bold; width: 25px;'>{display_val}</td>")
            html.append("<td style='width: 15px;'></td>")
        html.append("</tr>")
    html.append("</table></div>")
    return "".join(html)

def calculate_predictions(predictions_raw):
    prediction_results = {}
    for p in range(4):
        preds = predictions_raw[p]
        if not preds: continue
        
        angka_stats = {}
        for x in preds:
            v = x['val']
            l = x['length']
            if v not in angka_stats:
                angka_stats[v] = {'long_count': 0, 'short_count': 0, 'total': 0, 'max_len': 0}
            angka_stats[v]['total'] += 1
            
            if l >= 4: angka_stats[v]['long_count'] += 1
            else: angka_stats[v]['short_count'] += 1
                
            if l > angka_stats[v]['max_len']: angka_stats[v]['max_len'] = l

        kuat_candidates = []
        cadangan_candidates = []
        all_vals = [x['val'] for x in preds]
        
        for v, stats in angka_stats.items():
            if stats['long_count'] > 0 or stats['short_count'] > 1:
                kuat_candidates.append(v)
            else:
                cadangan_candidates.append(v)
        
        kuat_candidates.sort(key=lambda x: (angka_stats[x]['total'], angka_stats[x]['max_len']), reverse=True)
        cadangan_candidates.sort(key=lambda x: (angka_stats[x]['total'], angka_stats[x]['max_len']), reverse=True)
        
        angka_kuat = [kuat_candidates[0]] if kuat_candidates else []
        angka_cadangan = []
        if len(kuat_candidates) > 1: angka_cadangan = [kuat_candidates[1]]
        elif cadangan_candidates: angka_cadangan = [cadangan_candidates[0]]
            
        kuat_max_len = angka_stats[angka_kuat[0]]['max_len'] if angka_kuat else 0
        prediction_results[p] = {
            "kuat": angka_kuat, "cadangan": angka_cadangan,
            "max_len": kuat_max_len, "all_counts": Counter(all_vals)
        }
    return prediction_results

# --- MAIN APP ---
uploaded_file = st.file_uploader("Unggah Database Paito (.xlsx):", type=["xlsx"])

if uploaded_file:
    try:
        file_bytes = uploaded_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        
        hari_tabel = ["Sabtu", "Minggu", "Senin", "Selasa", "Rabu", "Kamis", "Jumat"]
        start_cols = [1, 6, 11, 16, 21, 26, 31]
        batas_bawah = ws.max_row 

        # Membuat Tabs untuk 2 Mode Pencarian
        tab1, tab2 = st.tabs(["🔍 Mode Scan Klasik (Mundur)", "⏭️ Mode Validasi Progresif (Maju)"])

        # ==========================================
        # TAB 1: MODE SCAN KLASIK
        # ==========================================
        with tab1:
            st.header("Mode Scan Klasik (Auto-Fetch Mundur)")
            
            c_opt1, c_opt2 = st.columns(2)
            with c_opt1: hari_terpilih = st.selectbox("Pilih Hari Utama (Hari Ini):", hari_tabel, index=4)
            with c_opt2: target_row_utama = st.number_input("Baris Target (Default: Baris Terakhir)", min_value=1, value=ws.max_row)

            idx_day0 = hari_tabel.index(hari_terpilih)
            inputs, update_targets = [], []
            cols = st.columns(6)
            
            current_r = target_row_utama
            for i in range(6):
                d_idx = (idx_day0 - i) % 7
                if i > 0:
                    prev_d_idx = (idx_day0 - (i - 1)) % 7
                    if prev_d_idx == 0 and d_idx == 6: current_r -= 1
                
                c_start = start_cols[d_idx]
                vals = [ws.cell(row=current_r, column=c_start+j).value for j in range(4)]
                if any(v is not None for v in vals):
                    auto_val = "".join([str(clean_int(v)) if clean_int(v) is not None else "0" for v in vals])
                else: auto_val = "0000"
                    
                with cols[i]:
                    user_val = st.text_input(f"{hari_tabel[d_idx]} (-{i}):", value=auto_val, max_chars=4, key=f"t1_{i}")
                    inputs.append(user_val)
                    update_targets.append((current_r, c_start))

            for i, user_val in enumerate(inputs):
                r_target, c_start = update_targets[i]
                user_val = user_val.ljust(4, '0')[:4] 
                for offset in range(4):
                    try: ws.cell(row=r_target, column=c_start + offset).value = int(user_val[offset])
                    except ValueError: pass

            st.divider()
            c1, c2, c3, c4 = st.columns(4)
            with c1: c_lurus = st.checkbox("Garis Lurus", value=True)
            with c2: c_naik = st.checkbox("Diagonal Naik", value=True)
            with c3: c_turun = st.checkbox("Diagonal Turun", value=True)
            with c4: c_terdekat = st.checkbox("Toleransi (+/- 1 & Indeks)", value=True)
            
            use_single_ref = st.checkbox("Mode Acuan Posisi Tunggal", value=False)
            ref_pos_name = st.selectbox("Posisi Acuan:", ["As", "Kop", "Kepala", "Ekor"], index=0, disabled=not use_single_ref)
            ref_pos_offset = ["As", "Kop", "Kepala", "Ekor"].index(ref_pos_name)

            if st.button("JALANKAN ANALISA KLASIK", type="primary"):
                cell_patterns = {}
                total_stats = {6: 0, 5: 0, 4: 0, 3: 0}
                days_indices = [(idx_day0 - k) % 7 for k in range(6)]
                predictions_raw = {0: [], 1: [], 2: [], 3: []}
                prediction_cells = set()

                for pos_offset in range(4):
                    current_allowed = []
                    for k in range(6):
                        val_str = inputs[k]
                        digit = int(val_str[ref_pos_offset if use_single_ref else pos_offset])
                        indek = (digit + 5) % 10
                        if c_terdekat:
                            allowed = {digit, indek, (digit-1)%10, (digit+1)%10, (indek-1)%10, (indek+1)%10}
                            current_allowed.append(list(allowed))
                        else: current_allowed.append([digit, indek])

                    for r_start in range(1, batas_bawah):
                        for mode in ["Lurus", "Naik", "Turun"]:
                            if (mode == "Lurus" and not c_lurus) or (mode == "Naik" and not c_naik) or (mode == "Turun" and not c_turun): continue
                            
                            for length in [6, 5, 4, 3]:
                                path, valid = [], True
                                for k in range(length):
                                    r_target = r_start if mode == "Lurus" else (r_start - k if mode == "Naik" else r_start + k)
                                    if r_target < 1 or r_target >= batas_bawah: valid = False; break
                                    
                                    val = clean_int(ws.cell(row=r_target, column=start_cols[days_indices[k]] + pos_offset).value)
                                    if val not in current_allowed[k]: valid = False; break
                                    path.append((r_target, start_cols[days_indices[k]] + pos_offset))
                                
                                if valid:
                                    total_stats[length] += 1
                                    for r_c, c_c in path: cell_patterns[(r_c, c_c)] = {"length": length, "pos": pos_offset}
                                    
                                    r_next = r_start if mode == "Lurus" else (r_start + 1 if mode == "Naik" else r_start - 1)
                                    if 1 <= r_next < batas_bawah:
                                        c_next = start_cols[(idx_day0 + 1) % 7] + pos_offset
                                        pred_val = clean_int(ws.cell(row=r_next, column=c_next).value)
                                        if pred_val is not None:
                                            predictions_raw[pos_offset].append({"val": pred_val, "length": length})
                                            prediction_cells.add((r_next, c_next))
                                    break 

                st.session_state.highlighted = cell_patterns
                st.session_state.stats = total_stats
                st.session_state.prediction_results = calculate_predictions(predictions_raw)
                st.session_state.prediction_cells = prediction_cells
                st.session_state.scanned = True

            if st.session_state.get("scanned"):
                st.divider()
                st.subheader("🎯 Prediksi Hari Berikutnya (Klasik)")
                pos_names = ["As", "Kop", "Kepala", "Ekor"]
                pred_cols = st.columns(4)
                for p in range(4):
                    with pred_cols[p]:
                        st.markdown(f"**Posisi {pos_names[p]}**")
                        if p in st.session_state.get("prediction_results", {}):
                            res = st.session_state.prediction_results[p]
                            kuat_str = str(res['kuat'][0]) if res['kuat'] else "-"
                            cadangan_str = str(res['cadangan'][0]) if res['cadangan'] else "-"
                            pola_info = f"*(Pola {res['max_len']} Baris)*" if res['kuat'] else "*(Tidak Ada)*"
                            st.success(f"🔥 **Kuat:** {kuat_str}\n\n{pola_info}")
                            st.info(f"🛡️ **Cadangan:** {cadangan_str}")
                        else: st.write("Belum ada pola")
                
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Pola 6 Hari", f"{st.session_state.stats[6]} Jalur")
                c2.metric("Pola 5 Hari", f"{st.session_state.stats[5]} Jalur")
                c3.metric("Pola 4 Hari", f"{st.session_state.stats[4]} Jalur")
                c4.metric("Pola 3 Hari", f"{st.session_state.stats[3]} Jalur")

                st.markdown(render_grid(ws, start_cols, hari_tabel, st.session_state.highlighted, st.session_state.prediction_cells), unsafe_allow_html=True)


        # ==========================================
        # TAB 2: MODE VALIDASI PROGRESIF
        # ==========================================
        with tab2:
            st.header("Mode Validasi Progresif (Forward Tracking)")
            st.info("💡 **Cara Penggunaan:** Isi mulai dari Langkah 1. Klik Lacak. Lalu isi Langkah 2 untuk memfilter pola yang nyambung, dan seterusnya sampai Langkah target.")
            
            pc1, pc2 = st.columns(2)
            with pc1: prog_hari_awal = st.selectbox("Mulai dari Hari (Langkah 1):", hari_tabel, index=2, key="prog_awal")
            with pc2: prog_jml_step = st.slider("Jumlah Hari Pembentuk Pola:", min_value=2, max_value=6, value=3, key="prog_jml")
            
            prog_idx_awal = hari_tabel.index(prog_hari_awal)
            prog_days_indices = [(prog_idx_awal + i) % 7 for i in range(prog_jml_step)]
            prog_pred_idx = (prog_idx_awal + prog_jml_step) % 7
            
            st.markdown(f"**Target Prediksi Output:** Hari **{hari_tabel[prog_pred_idx]}**")
            
            prog_inputs = []
            cols_prog = st.columns(prog_jml_step)
            for i in range(prog_jml_step):
                with cols_prog[i]:
                    val = st.text_input(f"Langkah {i+1} ({hari_tabel[prog_days_indices[i]]}):", max_chars=4, key=f"p_in_{i}")
                    prog_inputs.append(val)
                    
            st.divider()
            pc1, pc2, pc3, pc4 = st.columns(4)
            with pc1: p_lurus = st.checkbox("Garis Lurus", value=True, key="p_lurus")
            with pc2: p_naik = st.checkbox("Diagonal Naik", value=True, key="p_naik")
            with pc3: p_turun = st.checkbox("Diagonal Turun", value=True, key="p_turun")
            with pc4: p_terdekat = st.checkbox("Toleransi (+/- 1 & Indeks)", value=True, key="p_terdekat")

            if st.button("Lacak Pola Progresif", type="primary", key="btn_prog"):
                # Menghitung input yang diisi berurutan
                valid_steps = 0
                for v in prog_inputs:
                    if len(v.strip()) == 4: valid_steps += 1
                    elif len(v.strip()) > 0: 
                        st.error("Setiap langkah harus diisi tepat 4 digit angka!")
                        st.stop()
                    else: break
                    
                if valid_steps == 0:
                    st.warning("Silakan isi minimal Langkah 1!")
                else:
                    p_cell_patterns = {}
                    p_predictions_raw = {0:[], 1:[], 2:[], 3:[]}
                    p_prediction_cells = set()
                    found_count = 0
                    
                    for pos_offset in range(4):
                        current_allowed = []
                        for k in range(valid_steps):
                            val_str = prog_inputs[k]
                            digit = int(val_str[pos_offset])
                            indek = (digit + 5) % 10
                            if p_terdekat:
                                allowed = {digit, indek, (digit-1)%10, (digit+1)%10, (indek-1)%10, (indek+1)%10}
                                current_allowed.append(list(allowed))
                            else: current_allowed.append([digit, indek])
                                
                        for r_start in range(1, batas_bawah):
                            for mode in ["Lurus", "Naik", "Turun"]:
                                if (mode == "Lurus" and not p_lurus) or (mode == "Naik" and not p_naik) or (mode == "Turun" and not p_turun): continue
                                
                                path, valid = [], True
                                for k in range(valid_steps):
                                    r_target = r_start if mode == "Lurus" else (r_start - k if mode == "Naik" else r_start + k)
                                    if r_target < 1 or r_target >= batas_bawah: valid = False; break
                                        
                                    c_target = start_cols[prog_days_indices[k]] + pos_offset
                                    val = clean_int(ws.cell(row=r_target, column=c_target).value)
                                    if val not in current_allowed[k]: valid = False; break
                                    path.append((r_target, c_target))
                                    
                                if valid:
                                    found_count += 1
                                    for r_c, c_c in path:
                                        p_cell_patterns[(r_c, c_c)] = {"length": valid_steps, "pos": pos_offset}
                                        
                                    # Jika target sudah komplit semua langkah, cari prediksinya
                                    if valid_steps == prog_jml_step:
                                        r_next = r_start if mode == "Lurus" else (r_start - valid_steps if mode == "Naik" else r_start + valid_steps)
                                        if 1 <= r_next < batas_bawah:
                                            c_next = start_cols[prog_pred_idx] + pos_offset
                                            pred_val = clean_int(ws.cell(row=r_next, column=c_next).value)
                                            if pred_val is not None:
                                                p_predictions_raw[pos_offset].append({"val": pred_val, "length": valid_steps})
                                                p_prediction_cells.add((r_next, c_next))
                                                
                    st.session_state.prog_highlighted = p_cell_patterns
                    st.session_state.prog_predictions = calculate_predictions(p_predictions_raw) if valid_steps == prog_jml_step else None
                    st.session_state.prog_pred_cells = p_prediction_cells
                    st.session_state.prog_found = found_count
                    st.session_state.prog_valid_steps = valid_steps
                    st.session_state.prog_scanned = True
                    
            if st.session_state.get("prog_scanned"):
                v_steps = st.session_state.prog_valid_steps
                st.success(f"✔️ Ditemukan **{st.session_state.prog_found} jalur** yang saling menyambung untuk {v_steps} langkah yang diinput!")
                
                if v_steps == prog_jml_step and st.session_state.prog_predictions:
                    st.subheader(f"🎯 Hasil Prediksi Hari {hari_tabel[prog_pred_idx]}")
                    pos_names = ["As", "Kop", "Kepala", "Ekor"]
                    pred_cols = st.columns(4)
                    for p in range(4):
                        with pred_cols[p]:
                            st.markdown(f"**Posisi {pos_names[p]}**")
                            res = st.session_state.prog_predictions[p]
                            kuat_str = str(res['kuat'][0]) if res.get('kuat') else "-"
                            st.success(f"🔥 **Kuat:** {kuat_str}")
                
                st.markdown(render_grid(ws, start_cols, hari_tabel, st.session_state.prog_highlighted, st.session_state.prog_pred_cells), unsafe_allow_html=True)

    except Exception as e:
        st.error(f"Terjadi kesalahan: {e}")
